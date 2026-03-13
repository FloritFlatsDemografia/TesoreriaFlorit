import re
import unicodedata
from datetime import date
from io import BytesIO
from typing import Optional

import pandas as pd
import streamlit as st
import altair as alt
import openpyxl

# -----------------------------
# Config
# -----------------------------
st.set_page_config(page_title="APP Tesorería", layout="wide")
st.title("APP Tesorería — Dashboard")

# -----------------------------
# Helpers
# -----------------------------
MIN_REQUIRED = [
    "GENERAL",
    "TIPO",
    "DEPARTAMENTO",
    "NATURALEZA",
    "PERIODICIDAD",
    "REGLA_FECHA",
    "VALOR_FECHA",
    "LAG",
    "AJUSTE FINDE",
]

OPTIONAL_NEW_COLUMNS = [
    "RAIZ CUENTA CONTABLE",
    "EMPRESA",
    "CLIENTE",
    "FECHA CARGO GASTO",
    "PERIODO_SERVICIO",
    "IVA_EN_FACTURA",
    "IVA_%",
    "IVA_SENTIDO",
    "TRATAMIENTO_IVA",
]

def strip_accents(text: str) -> str:
    text = str(text)
    return "".join(
        c for c in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(c)
    )

def normalize_colname(name: str) -> str:
    s = strip_accents(str(name))
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip().upper()
    return s

def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [normalize_colname(c) for c in df.columns]
    return df

def apply_column_aliases(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    alias_map = {
        "PREVISION": "IMPORTE PRONOSTICADO",
        "IMPORTE_PRONOSTICADO": "IMPORTE PRONOSTICADO",
        "IMPORTE_REAL": "IMPORTE REAL",
        "FECHA CARGO GASTO": "VALOR_FECHA",
        "FECHA_CARGO_GASTO": "VALOR_FECHA",
        "AJUSTE_FINDE": "AJUSTE FINDE",
        "REGLA FECHA": "REGLA_FECHA",
        "VALOR FECHA": "VALOR_FECHA",
        "TRATAMINETO_IVA": "TRATAMIENTO_IVA",
        "TRATAMIENTO IVA": "TRATAMIENTO_IVA",
        "RAIZ_CUENTA_CONTABLE": "RAIZ CUENTA CONTABLE",
        "PERIODO SERVICIO": "PERIODO_SERVICIO",
    }

    rename_dict = {}
    for col in df.columns:
        if col in alias_map:
            target = alias_map[col]
            if target not in df.columns:
                rename_dict[col] = target

    if rename_dict:
        df = df.rename(columns=rename_dict)

    return df

def find_header_row(df_raw: pd.DataFrame) -> int | None:
    for i in range(min(80, len(df_raw))):
        row = df_raw.iloc[i].astype(str).map(normalize_colname).tolist()
        if "GENERAL" in row and "TIPO" in row:
            return i
    return None

def is_pagado(v) -> bool:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    s = str(v).strip().lower()
    return s in {"✓", "✅", "x", "si", "sí", "true", "1", "ok", "pagado", "y", "yes"}

def eur(x):
    try:
        return f"{float(x):,.2f} €"
    except Exception:
        return ""

def color_saldo(v):
    try:
        v = float(v)
    except Exception:
        return ""
    if v > 0:
        return "color: green; font-weight: 700;"
    if v < 0:
        return "color: red; font-weight: 700;"
    return ""

def estado_cobro_pago(tipo: str, pagado_bool: bool) -> str:
    t = (tipo or "").upper().strip()
    if t == "INGRESO":
        return "COBRADO" if pagado_bool else "PENDIENTE"
    if t == "GASTO":
        return "PAGADO" if pagado_bool else "PENDIENTE"
    return ""

def style_estado(val: str):
    s = str(val).strip().lower()
    if s == "pendiente":
        return "background-color: #FDE6C8; font-weight: 700;"
    if s == "cobrado":
        return "background-color: #D8F3DC; color: #0B6E2E; font-weight: 700;"
    if s == "pagado":
        return "background-color: #F8D7DA; color: #8A1C1C; font-weight: 700;"
    return ""

def concept_has_explicit_month(concepto: str) -> bool:
    s = strip_accents(str(concepto)).upper()
    meses = [
        "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
        "JULIO", "AGOSTO", "SEPTIEMBRE", "SETIEMBRE",
        "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
    ]
    return any(m in s for m in meses)

# -----------------------------
# Leer hoja BANCOS
# -----------------------------
def read_bancos_from_excel(uploaded_file) -> dict:
    uploaded_file.seek(0)
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if str(name).strip().upper() == "BANCOS":
            sheet_name = name
            break
    if sheet_name is None:
        raise ValueError("No encuentro la hoja 'BANCOS' en el Excel.")

    ws = wb[sheet_name]

    mapping = {}
    for r in range(1, ws.max_row + 1):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if k is None:
            continue
        mapping[str(k).strip().upper()] = v

    def get_num(key: str):
        v = mapping.get(key)
        if v is None:
            return None
        try:
            return float(v)
        except Exception:
            return None

    total_bancos = get_num("TOTAL BANCOS")
    suplidos = get_num("CUENTA SUPLIDOS")
    efectivo = get_num("CUENTA DE EFECTIVO")

    if total_bancos is None:
        raise ValueError("En hoja BANCOS no puedo leer un número en 'TOTAL BANCOS' (columna €).")

    return {
        "total_bancos": total_bancos,
        "cuenta_suplidos": suplidos,
        "cuenta_efectivo": efectivo
    }

# -----------------------------
# Leer catálogo
# -----------------------------
def read_catalog_from_excel(uploaded_file) -> pd.DataFrame:
    uploaded_file.seek(0)
    raw = pd.read_excel(uploaded_file, sheet_name=0, header=None, engine="openpyxl")
    header_idx = find_header_row(raw)
    if header_idx is None:
        raise ValueError("No encuentro la fila de cabecera (debe contener 'GENERAL' y 'TIPO').")

    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name=0, header=header_idx, engine="openpyxl")
    df = normalize_cols(df)
    df = apply_column_aliases(df)

    missing = [c for c in MIN_REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(
            f"Faltan columnas requeridas: {missing}. "
            f"Columnas detectadas: {list(df.columns)}"
        )

    for c in OPTIONAL_NEW_COLUMNS:
        if c not in df.columns:
            df[c] = ""

    if "PRORRATEO" not in df.columns:
        df["PRORRATEO"] = ""
    df["PRORRATEO"] = df["PRORRATEO"].astype(str).str.strip().str.upper()

    if "PAGADO" not in df.columns:
        df["PAGADO"] = ""

    if "FECHA" in df.columns:
        df["FECHA_PAGO"] = pd.to_datetime(df["FECHA"], errors="coerce").dt.normalize()
    else:
        df["FECHA_PAGO"] = pd.NaT

    df["PAGADO_BOOL"] = df["PAGADO"].apply(is_pagado)

    if "IMPORTE PRONOSTICADO" not in df.columns:
        raise ValueError("Falta columna: 'IMPORTE PRONOSTICADO' o su alias 'PREVISION'")

    if "IMPORTE REAL" not in df.columns:
        df["IMPORTE REAL"] = 0.0

    if "HASTA" not in df.columns:
        df["HASTA"] = pd.NaT
    df["HASTA"] = pd.to_datetime(df["HASTA"], errors="coerce").dt.normalize()

    df = df.dropna(how="all").copy()

    for c in [
        "GENERAL",
        "TIPO",
        "DEPARTAMENTO",
        "NATURALEZA",
        "PERIODICIDAD",
        "REGLA_FECHA",
        "AJUSTE FINDE",
        "CLIENTE",
    ]:
        if c not in df.columns:
            df[c] = ""
        df[c] = df[c].astype(str).str.strip()

    df["TIPO"] = df["TIPO"].str.upper()
    df["DEPARTAMENTO"] = df["DEPARTAMENTO"].str.upper()
    df["NATURALEZA"] = df["NATURALEZA"].str.upper()
    df["PERIODICIDAD"] = df["PERIODICIDAD"].str.upper()
    df["REGLA_FECHA"] = df["REGLA_FECHA"].str.upper()
    df["AJUSTE FINDE"] = df["AJUSTE FINDE"].str.upper()

    df["IMPORTE_PRON"] = pd.to_numeric(df["IMPORTE PRONOSTICADO"], errors="coerce").fillna(0.0)
    df["IMPORTE_REAL"] = pd.to_numeric(df["IMPORTE REAL"], errors="coerce").fillna(0.0)
    df["LAG"] = pd.to_numeric(df["LAG"], errors="coerce").fillna(0).astype(int)

    def to_day_of_month(v):
        if pd.isna(v):
            return None
        if isinstance(v, pd.Timestamp):
            return int(v.day)
        if hasattr(v, "day"):
            try:
                return int(v.day)
            except Exception:
                pass

        s = str(v).strip()

        if re.fullmatch(r"\d{1,2}", s):
            d = int(s)
            if 1 <= d <= 31:
                return d

        try:
            dt = pd.to_datetime(s, errors="coerce")
            if pd.notna(dt):
                return int(dt.day)
        except Exception:
            pass

        return None

    df["DIA_MES"] = df["VALOR_FECHA"].apply(to_day_of_month)
    df["FECHA_FIJA"] = pd.to_datetime(df["VALOR_FECHA"], errors="coerce").dt.normalize()

    return df

def next_business_day(d: pd.Timestamp) -> pd.Timestamp:
    if d.weekday() == 5:
        return d + pd.Timedelta(days=2)
    if d.weekday() == 6:
        return d + pd.Timedelta(days=1)
    return d

def previous_business_day(d: pd.Timestamp) -> pd.Timestamp:
    if d.weekday() == 5:
        return d - pd.Timedelta(days=1)
    if d.weekday() == 6:
        return d - pd.Timedelta(days=2)
    return d

def add_business_days(d: pd.Timestamp, n: int) -> pd.Timestamp:
    cur = d
    step = 1 if n >= 0 else -1
    remaining = abs(n)
    while remaining > 0:
        cur = cur + pd.Timedelta(days=step)
        if cur.weekday() < 5:
            remaining -= 1
    return cur

def months_step_from_periodicidad(periodicidad: str) -> int:
    p = (periodicidad or "").upper().strip()
    if p == "MENSUAL":
        return 1
    if p in ("BIMESTRAL", "BIMENSUAL"):
        return 2
    if p == "TRIMESTRAL":
        return 3
    if p == "SEMESTRAL":
        return 6
    return 1

def resolve_base_date_for_row(r: pd.Series, reference_start: pd.Timestamp) -> Optional[pd.Timestamp]:
    periodicidad = str(r.get("PERIODICIDAD", "")).upper().strip()
    regla = str(r.get("REGLA_FECHA", "")).upper().strip()

    if periodicidad in ("PUNTUAL", "ONE-OFF", "ONEOFF", "ANUAL", "SEMANAL"):
        if pd.notna(r.get("FECHA_FIJA")):
            return pd.Timestamp(r["FECHA_FIJA"]).normalize()
        return pd.NaT

    if periodicidad in ("MENSUAL", "BIMESTRAL", "BIMENSUAL", "TRIMESTRAL", "SEMESTRAL"):
        if pd.notna(r.get("FECHA_FIJA")):
            return pd.Timestamp(r["FECHA_FIJA"]).normalize()

        day = r.get("DIA_MES")
        if day and not pd.isna(day):
            y0, m0 = reference_start.year, reference_start.month
            last_day0 = (pd.Timestamp(year=y0, month=m0, day=1) + pd.offsets.MonthEnd(0)).day
            return pd.Timestamp(year=y0, month=m0, day=min(int(day), int(last_day0))).normalize()

        if regla == "ULTIMO_HABIL":
            y0, m0 = reference_start.year, reference_start.month
            d = (pd.Timestamp(year=y0, month=m0, day=1) + pd.offsets.MonthEnd(0)).normalize()
            return previous_business_day(d).normalize()

    return pd.NaT

def apply_row_adjustments(d: pd.Timestamp, ajuste: str, lag: int) -> pd.Timestamp:
    if pd.isna(d):
        return d

    d = pd.Timestamp(d).normalize()

    if ajuste == "SIG_HABIL":
        d = next_business_day(d)
    elif ajuste in {"ANT_HABIL", "PREV_HABIL"}:
        d = previous_business_day(d)

    if lag:
        d = add_business_days(d, lag)

    return d.normalize()

def generate_events_from_catalog(catalog: pd.DataFrame, start_date: pd.Timestamp, months_horizon: int) -> pd.DataFrame:
    end_date = (start_date + pd.offsets.MonthBegin(months_horizon + 1)).normalize()
    rows = []

    for _, r in catalog.iterrows():
        periodicidad = str(r.get("PERIODICIDAD", "")).upper().strip()
        regla = str(r.get("REGLA_FECHA", "")).upper().strip()
        ajuste = str(r.get("AJUSTE FINDE", "")).upper().strip()
        lag = int(r.get("LAG", 0))
        prorrateo = str(r.get("PRORRATEO", "")).upper().strip()
        concepto = str(r.get("GENERAL", "")).strip()
        force_single_event = concept_has_explicit_month(concepto)

        hasta = r.get("HASTA", pd.NaT)
        hasta = pd.Timestamp(hasta).normalize() if not pd.isna(hasta) else pd.NaT

        def apply_adjustments(d: pd.Timestamp) -> pd.Timestamp:
            return apply_row_adjustments(d, ajuste, lag)

        def within_limits(d: pd.Timestamp) -> bool:
            if d < start_date or d > end_date:
                return False
            if not pd.isna(hasta) and d > hasta:
                return False
            return True

        def add_one(d_base: pd.Timestamp):
            d_base = pd.Timestamp(d_base).normalize()
            if not within_limits(d_base):
                return
            d_adj = apply_adjustments(d_base)
            if d_adj < start_date or d_adj > end_date:
                return
            rows.append((d_adj, r))

        def add_prorrateo_diario_for_month(d_base: pd.Timestamp):
            d_base = pd.Timestamp(d_base).normalize()
            y, m = d_base.year, d_base.month
            month_start = pd.Timestamp(year=y, month=m, day=1).normalize()
            month_end = (month_start + pd.offsets.MonthEnd(0)).normalize()
            days_in_month = int((month_end - month_start).days) + 1
            if days_in_month <= 0:
                return

            imp_pron_day = float(r.get("IMPORTE_PRON", 0.0)) / days_in_month
            imp_real_day = float(r.get("IMPORTE_REAL", 0.0)) / days_in_month

            d = month_start
            while d <= month_end:
                if not within_limits(d):
                    d += pd.Timedelta(days=1)
                    continue
                rr = r.copy()
                rr["IMPORTE_PRON"] = imp_pron_day
                rr["IMPORTE_REAL"] = imp_real_day
                rows.append((d.normalize(), rr))
                d += pd.Timedelta(days=1)

        if force_single_event:
            fecha_base = resolve_base_date_for_row(r, start_date)
            if pd.notna(fecha_base):
                add_one(fecha_base)
            continue

        if periodicidad in ("PUNTUAL", "ONE-OFF", "ONEOFF"):
            if pd.isna(r.get("FECHA_FIJA")):
                continue
            add_one(pd.Timestamp(r["FECHA_FIJA"]).normalize())
            continue

        if periodicidad == "ANUAL":
            if pd.isna(r.get("FECHA_FIJA")):
                continue
            base = pd.Timestamp(r["FECHA_FIJA"]).normalize()
            year = start_date.year
            while True:
                try:
                    candidate = pd.Timestamp(year=year, month=base.month, day=base.day)
                except Exception:
                    break
                if candidate > end_date:
                    break
                add_one(candidate.normalize())
                year += 1
            continue

        if periodicidad == "SEMANAL":
            anchor = r.get("FECHA_FIJA")
            if pd.isna(anchor):
                continue
            anchor = pd.Timestamp(anchor).normalize()

            if not pd.isna(hasta):
                stop = min(hasta, end_date)
            else:
                month_start = anchor.replace(day=1)
                stop = (month_start + pd.offsets.MonthEnd(0)).normalize()
                stop = min(stop, end_date)

            d = anchor
            while d <= stop:
                add_one(d.normalize())
                d = d + pd.Timedelta(days=7)
            continue

        if periodicidad in ("MENSUAL", "BIMESTRAL", "BIMENSUAL", "TRIMESTRAL", "SEMESTRAL"):
            step = months_step_from_periodicidad(periodicidad)

            if not pd.isna(r.get("FECHA_FIJA")):
                base_date = pd.Timestamp(r["FECHA_FIJA"]).normalize()
                anchor_day = base_date.day
            else:
                day = r.get("DIA_MES")
                if not day or pd.isna(day):
                    continue
                y0, m0 = start_date.year, start_date.month
                last_day0 = (pd.Timestamp(year=y0, month=m0, day=1) + pd.offsets.MonthEnd(0)).day
                base_date = pd.Timestamp(year=y0, month=m0, day=min(int(day), int(last_day0))).normalize()
                anchor_day = base_date.day

            current = base_date

            while current <= end_date:
                y, m = current.year, current.month

                if regla in ("DIA_MES", "FECHA_FIJA"):
                    last_day = (pd.Timestamp(year=y, month=m, day=1) + pd.offsets.MonthEnd(0)).day
                    d_base = pd.Timestamp(year=y, month=m, day=min(int(anchor_day), int(last_day))).normalize()
                elif regla == "ULTIMO_HABIL":
                    d_base = (pd.Timestamp(year=y, month=m, day=1) + pd.offsets.MonthEnd(0)).normalize()
                    d_base = previous_business_day(d_base).normalize()
                else:
                    d_base = None

                if d_base is not None:
                    if prorrateo == "DIARIO":
                        if within_limits(d_base):
                            add_prorrateo_diario_for_month(d_base)
                    else:
                        add_one(d_base)

                current = (current + pd.DateOffset(months=step)).normalize()

    if not rows:
        return pd.DataFrame(columns=[
            "FECHA", "CONCEPTO", "TIPO", "DEPARTAMENTO", "CLIENTE", "IMPORTE_PRON", "IMPORTE_REAL",
            "NATURALEZA", "PAGADO_BOOL", "FECHA_PAGO", "PRORRATEO", "ESTATUS"
        ])

    out = pd.DataFrame([{
        "FECHA": d,
        "CONCEPTO": rr["GENERAL"],
        "TIPO": rr["TIPO"],
        "DEPARTAMENTO": rr["DEPARTAMENTO"],
        "CLIENTE": rr.get("CLIENTE", ""),
        "IMPORTE_PRON": float(rr.get("IMPORTE_PRON", 0.0)),
        "IMPORTE_REAL": float(rr.get("IMPORTE_REAL", 0.0)),
        "NATURALEZA": rr.get("NATURALEZA", ""),
        "PAGADO_BOOL": bool(rr.get("PAGADO_BOOL", False)),
        "FECHA_PAGO": rr.get("FECHA_PAGO", pd.NaT),
        "PRORRATEO": str(rr.get("PRORRATEO", "")).upper().strip(),
    } for d, rr in rows])

    out["FECHA_PAGO"] = pd.to_datetime(out["FECHA_PAGO"], errors="coerce").dt.normalize()
    out["ESTATUS"] = out.apply(
        lambda x: estado_cobro_pago(x.get("TIPO", ""), bool(x.get("PAGADO_BOOL", False))),
        axis=1
    )
    return out.sort_values("FECHA").reset_index(drop=True)

def build_real_events_from_catalog(catalog: pd.DataFrame, start_date: pd.Timestamp, months_horizon: int) -> pd.DataFrame:
    end_date = (start_date + pd.offsets.MonthBegin(months_horizon + 1)).normalize()
    rows = []

    for _, r in catalog.iterrows():
        pagado_bool = bool(r.get("PAGADO_BOOL", False))
        importe_real = float(r.get("IMPORTE_REAL", 0.0))
        fecha_pago = r.get("FECHA_PAGO", pd.NaT)

        tiene_senal_real = pagado_bool or (abs(importe_real) > 0) or pd.notna(fecha_pago)
        if not tiene_senal_real:
            continue

        fecha_base = resolve_base_date_for_row(r, start_date)
        ajuste = str(r.get("AJUSTE FINDE", "")).upper().strip()
        lag = int(r.get("LAG", 0))

        if pd.notna(fecha_pago):
            fecha_real = pd.Timestamp(fecha_pago).normalize()
        elif pd.notna(fecha_base):
            fecha_real = apply_row_adjustments(fecha_base, ajuste, lag)
        else:
            continue

        hasta = r.get("HASTA", pd.NaT)
        hasta = pd.Timestamp(hasta).normalize() if not pd.isna(hasta) else pd.NaT

        if fecha_real < start_date or fecha_real > end_date:
            continue
        if not pd.isna(hasta) and fecha_real > hasta:
            continue

        rows.append({
            "FECHA": fecha_real,
            "CONCEPTO": r.get("GENERAL", ""),
            "TIPO": r.get("TIPO", ""),
            "DEPARTAMENTO": r.get("DEPARTAMENTO", ""),
            "CLIENTE": r.get("CLIENTE", ""),
            "IMPORTE_PRON": float(r.get("IMPORTE_PRON", 0.0)),
            "IMPORTE_REAL": importe_real,
            "NATURALEZA": r.get("NATURALEZA", ""),
            "PAGADO_BOOL": pagado_bool,
            "FECHA_PAGO": pd.Timestamp(fecha_pago).normalize() if pd.notna(fecha_pago) else pd.NaT,
            "PRORRATEO": str(r.get("PRORRATEO", "")).upper().strip(),
            "ESTATUS": estado_cobro_pago(r.get("TIPO", ""), pagado_bool),
        })

    if not rows:
        return pd.DataFrame(columns=[
            "FECHA", "CONCEPTO", "TIPO", "DEPARTAMENTO", "CLIENTE", "IMPORTE_PRON", "IMPORTE_REAL",
            "NATURALEZA", "PAGADO_BOOL", "FECHA_PAGO", "PRORRATEO", "ESTATUS"
        ])

    out = pd.DataFrame(rows)
    out["FECHA"] = pd.to_datetime(out["FECHA"], errors="coerce").dt.normalize()
    out["FECHA_PAGO"] = pd.to_datetime(out["FECHA_PAGO"], errors="coerce").dt.normalize()
    out = out.sort_values(["FECHA", "CONCEPTO", "CLIENTE"]).reset_index(drop=True)
    return out

def compute_balance_from_amount(df: pd.DataFrame, starting_balance: float, amount_col: str) -> pd.DataFrame:
    df = df.copy()
    df["COBROS"] = df.apply(lambda x: x[amount_col] if x["TIPO"] == "INGRESO" else 0.0, axis=1)
    df["PAGOS"] = df.apply(lambda x: x[amount_col] if x["TIPO"] == "GASTO" else 0.0, axis=1)
    df["NETO"] = df["COBROS"] - df["PAGOS"]
    df["SALDO"] = starting_balance + df["NETO"].cumsum()
    return df

st.sidebar.header("Inputs")
saldo_fecha = st.sidebar.date_input("Fecha del saldo (hoy)", value=date.today())
months_horizon = st.sidebar.slider("Horizonte forecast (meses)", min_value=1, max_value=36, value=12)
dedupe_exact = st.sidebar.checkbox("Eliminar duplicados exactos (red de seguridad)", value=True)
uploaded = st.sidebar.file_uploader("Sube el Excel de catálogo (xlsx)", type=["xlsx"])

if not uploaded:
    st.info("Sube tu Excel para generar el dashboard.")
    st.stop()

try:
    bancos = read_bancos_from_excel(uploaded)
except Exception as e:
    st.error(f"Error leyendo hoja BANCOS: {e}")
    st.stop()

saldo_hoy = float(bancos["total_bancos"])
cuenta_suplidos = bancos.get("cuenta_suplidos", None)
cuenta_efectivo = bancos.get("cuenta_efectivo", None)

st.sidebar.header("Bancos (desde Excel)")
st.sidebar.metric("TOTAL BANCOS (saldo inicial)", eur(saldo_hoy))
if cuenta_suplidos is not None:
    st.sidebar.metric("CUENTA SUPLIDOS", eur(cuenta_suplidos))
if cuenta_efectivo is not None:
    st.sidebar.metric("CUENTA DE EFECTIVO", eur(cuenta_efectivo))

try:
    catalog = read_catalog_from_excel(uploaded)
except Exception as e:
    st.error(f"Error leyendo catálogo: {e}")
    st.stop()

start_ts = pd.Timestamp(saldo_fecha).normalize()

generated_pron = generate_events_from_catalog(
    catalog=catalog,
    start_date=start_ts,
    months_horizon=months_horizon
)

generated_real = build_real_events_from_catalog(
    catalog=catalog,
    start_date=start_ts,
    months_horizon=months_horizon
)

if dedupe_exact and not generated_pron.empty:
    generated_pron = generated_pron.drop_duplicates(
        subset=[
            "FECHA", "CONCEPTO", "TIPO", "DEPARTAMENTO", "CLIENTE",
            "IMPORTE_PRON", "IMPORTE_REAL", "PAGADO_BOOL", "FECHA_PAGO", "PRORRATEO"
        ],
        keep="first"
    )

if dedupe_exact and not generated_real.empty:
    generated_real = generated_real.drop_duplicates(
        subset=[
            "FECHA", "CONCEPTO", "TIPO", "DEPARTAMENTO", "CLIENTE",
            "IMPORTE_REAL", "PAGADO_BOOL", "FECHA_PAGO"
        ],
        keep="first"
    )

if generated_pron.empty and generated_real.empty:
    st.warning("No se generaron movimientos.")
    st.dataframe(catalog.head(50), use_container_width=True)
    st.stop()

st.sidebar.header("Filtros base")

all_deptos = sorted(
    pd.concat([
        generated_pron["DEPARTAMENTO"] if not generated_pron.empty else pd.Series(dtype=str),
        generated_real["DEPARTAMENTO"] if not generated_real.empty else pd.Series(dtype=str),
    ], ignore_index=True).dropna().astype(str).unique().tolist()
)

all_tipos = sorted(
    pd.concat([
        generated_pron["TIPO"] if not generated_pron.empty else pd.Series(dtype=str),
        generated_real["TIPO"] if not generated_real.empty else pd.Series(dtype=str),
    ], ignore_index=True).dropna().astype(str).unique().tolist()
)

sel_deptos = st.sidebar.multiselect("Departamento", options=all_deptos, default=all_deptos)
sel_tipos = st.sidebar.multiselect("Tipo", options=all_tipos, default=all_tipos)

base_filtered_pron = generated_pron[
    generated_pron["DEPARTAMENTO"].isin(sel_deptos) &
    generated_pron["TIPO"].isin(sel_tipos)
].copy().sort_values("FECHA").reset_index(drop=True) if not generated_pron.empty else generated_pron.copy()

base_filtered_real = generated_real[
    generated_real["DEPARTAMENTO"].isin(sel_deptos) &
    generated_real["TIPO"].isin(sel_tipos)
].copy().sort_values("FECHA").reset_index(drop=True) if not generated_real.empty else generated_real.copy()

pron_df = base_filtered_pron[~base_filtered_pron["PAGADO_BOOL"]].copy().sort_values("FECHA").reset_index(drop=True)

real_df = base_filtered_real.copy()
real_df["FECHA"] = pd.to_datetime(real_df["FECHA"], errors="coerce").dt.normalize()
real_df["ESTATUS"] = real_df.apply(
    lambda r: estado_cobro_pago(r.get("TIPO", ""), bool(r.get("PAGADO_BOOL", False))),
    axis=1
)
real_df = real_df.sort_values("FECHA").reset_index(drop=True)

consolidado_pron = compute_balance_from_amount(pron_df, saldo_hoy, "IMPORTE_PRON") if not pron_df.empty else pron_df.copy()
consolidado_real = compute_balance_from_amount(real_df, saldo_hoy, "IMPORTE_REAL") if not real_df.empty else real_df.copy()

if not consolidado_pron.empty:
    consolidado_pron["ESTATUS"] = consolidado_pron.apply(
        lambda r: estado_cobro_pago(r.get("TIPO", ""), bool(r.get("PAGADO_BOOL", False))),
        axis=1
    )

if not consolidado_real.empty:
    consolidado_real["ESTATUS"] = consolidado_real.apply(
        lambda r: estado_cobro_pago(r.get("TIPO", ""), bool(r.get("PAGADO_BOOL", False))),
        axis=1
    )

base_row = pd.DataFrame([{
    "FECHA": start_ts,
    "CONCEPTO": "SALDO BANCOS TOTAL",
    "TIPO": "SALDO",
    "DEPARTAMENTO": "",
    "CLIENTE": "",
    "IMPORTE_PRON": 0.0,
    "IMPORTE_REAL": 0.0,
    "NATURALEZA": "SALDO",
    "PAGADO_BOOL": False,
    "FECHA_PAGO": pd.NaT,
    "PRORRATEO": "",
    "ESTATUS": "",
    "COBROS": 0.0,
    "PAGOS": 0.0,
    "NETO": 0.0,
    "SALDO": saldo_hoy
}])

consolidado_pron2 = pd.concat([base_row, consolidado_pron], ignore_index=True) if not consolidado_pron.empty else base_row.copy()
consolidado_real2 = pd.concat([base_row, consolidado_real], ignore_index=True) if not consolidado_real.empty else base_row.copy()

st.sidebar.header("Búsqueda y rango (solo visualización)")
q_concepto = st.sidebar.text_input("Buscar concepto", value="").strip()
q_cliente = st.sidebar.text_input("Buscar cliente", value="").strip()

estatus_options = ["PAGADO", "COBRADO", "PENDIENTE"]
sel_estatus = st.sidebar.multiselect(
    "Estatus",
    options=estatus_options,
    default=estatus_options
)

min_d = min(consolidado_pron2["FECHA"].min(), consolidado_real2["FECHA"].min()).date()
max_d = max(consolidado_pron2["FECHA"].max(), consolidado_real2["FECHA"].max()).date()

date_range = st.sidebar.date_input(
    "Rango de fechas",
    value=(min_d, max_d),
    min_value=min_d,
    max_value=max_d
)

if isinstance(date_range, tuple) and len(date_range) == 2:
    d_from, d_to = date_range
else:
    d_from, d_to = min_d, max_d

def apply_visual_filters(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out = out[(out["FECHA"].dt.date >= d_from) & (out["FECHA"].dt.date <= d_to)].copy()

    if q_concepto:
        out = out[out["CONCEPTO"].astype(str).str.contains(q_concepto, case=False, na=False)].copy()

    if q_cliente:
        if "CLIENTE" in out.columns:
            out = out[out["CLIENTE"].astype(str).str.contains(q_cliente, case=False, na=False)].copy()
        else:
            out = out.iloc[0:0].copy()

    if sel_estatus:
        if "ESTATUS" in out.columns:
            out = out[out["ESTATUS"].astype(str).str.upper().isin(sel_estatus)].copy()
        else:
            out = out.iloc[0:0].copy()

    return out.sort_values("FECHA").reset_index(drop=True)

view_pron = apply_visual_filters(consolidado_pron2)
view_real = apply_visual_filters(consolidado_real2)

# -----------------------------
# KPIs
# -----------------------------
c1, c2, c3, c4, c5 = st.columns(5)
with c1:
    st.metric("Saldo inicial (TOTAL BANCOS)", eur(saldo_hoy))
with c2:
    st.metric("Neto periodo (PRON, pendientes)", eur(consolidado_pron["NETO"].sum() if not consolidado_pron.empty else 0.0))
with c3:
    saldo_pron_final = consolidado_pron["SALDO"].iloc[-1] if not consolidado_pron.empty else saldo_hoy
    st.metric("Saldo final (PRON, pendientes)", eur(saldo_pron_final))
with c4:
    saldo_real_final = consolidado_real["SALDO"].iloc[-1] if not consolidado_real.empty else saldo_hoy
    st.metric("Saldo final (REAL, estimado/ejecutado)", eur(saldo_real_final))
with c5:
    desviacion_total = saldo_real_final - saldo_pron_final
    st.metric("Desviación PRON vs REAL", eur(desviacion_total))

# -----------------------------
# Gráfico diario
# -----------------------------
st.subheader("Evolución de saldo — diario (Pronosticado vs Real)")

d_pron = consolidado_pron2[["FECHA", "SALDO"]].copy()
d_pron["FECHA"] = pd.to_datetime(d_pron["FECHA"]).dt.normalize()
d_pron = d_pron.groupby("FECHA", as_index=False)["SALDO"].last().rename(columns={"SALDO": "SALDO_PRON"})

d_real = consolidado_real2[["FECHA", "SALDO"]].copy()
d_real["FECHA"] = pd.to_datetime(d_real["FECHA"]).dt.normalize()
d_real = d_real.groupby("FECHA", as_index=False)["SALDO"].last().rename(columns={"SALDO": "SALDO_REAL"})

daily = pd.merge(d_pron, d_real, on="FECHA", how="outer").sort_values("FECHA")

all_days = pd.date_range(start=daily["FECHA"].min(), end=daily["FECHA"].max(), freq="D")
daily = daily.set_index("FECHA").reindex(all_days).rename_axis("FECHA").reset_index()
daily["SALDO_PRON"] = daily["SALDO_PRON"].ffill().fillna(saldo_hoy)
daily["SALDO_REAL"] = daily["SALDO_REAL"].ffill().fillna(saldo_hoy)

daily["DESVIACION"] = daily["SALDO_REAL"] - daily["SALDO_PRON"]
daily["DESV_COLOR"] = daily["DESVIACION"].apply(lambda x: "Desviación positiva" if x >= 0 else "Desviación negativa")

if cuenta_suplidos is not None:
    daily["LINEA_SUPLIDOS"] = float(cuenta_suplidos)
if cuenta_efectivo is not None:
    daily["LINEA_EFECTIVO"] = float(cuenta_efectivo)

zoom_start = pd.Timestamp(d_from)
zoom_end = pd.Timestamp(d_to)
daily_zoom = daily[(daily["FECHA"] >= zoom_start) & (daily["FECHA"] <= zoom_end)].copy()

# Líneas principales
value_vars = ["SALDO_PRON", "SALDO_REAL"]
series_map = {
    "SALDO_PRON": "Pronosticado (pendiente)",
    "SALDO_REAL": "Real (estimado/ejecutado)",
}

if "LINEA_EFECTIVO" in daily_zoom.columns:
    value_vars.append("LINEA_EFECTIVO")
    series_map["LINEA_EFECTIVO"] = "Cuenta efectivo"
if "LINEA_SUPLIDOS" in daily_zoom.columns:
    value_vars.append("LINEA_SUPLIDOS")
    series_map["LINEA_SUPLIDOS"] = "Cuenta suplidos"

plot_df = daily_zoom.melt(
    id_vars=["FECHA"],
    value_vars=value_vars,
    var_name="SERIE",
    value_name="SALDO"
)
plot_df["SERIE"] = plot_df["SERIE"].map(series_map)

domain = [
    "Pronosticado (pendiente)",
    "Real (estimado/ejecutado)",
    "Cuenta efectivo",
    "Cuenta suplidos"
]
range_ = [
    "#6BAED6",  # PRON
    "#08519C",  # REAL
    "#FF7F0E",  # EFECTIVO
    "#8A2BE2",  # SUPLIDOS violeta
]

base_chart = (
    alt.Chart(plot_df)
    .mark_line()
    .encode(
        x=alt.X("FECHA:T", title="Fecha"),
        y=alt.Y("SALDO:Q", title="Saldo"),
        color=alt.Color("SERIE:N", title="", scale=alt.Scale(domain=domain, range=range_)),
        tooltip=[
            alt.Tooltip("FECHA:T", title="Fecha"),
            alt.Tooltip("SERIE:N", title="Serie"),
            alt.Tooltip("SALDO:Q", title="Saldo", format=",.2f"),
        ],
    )
)

# Línea de desviación con color dinámico
desv_chart = (
    alt.Chart(daily_zoom)
    .mark_line(strokeWidth=3)
    .encode(
        x=alt.X("FECHA:T", title="Fecha"),
        y=alt.Y("DESVIACION:Q", title="Saldo"),
        color=alt.Color(
            "DESV_COLOR:N",
            title="",
            scale=alt.Scale(
                domain=["Desviación positiva", "Desviación negativa"],
                range=["#1A9850", "#D73027"]
            )
        ),
        tooltip=[
            alt.Tooltip("FECHA:T", title="Fecha"),
            alt.Tooltip("DESVIACION:Q", title="Desviación", format=",.2f"),
            alt.Tooltip("DESV_COLOR:N", title="Tipo"),
        ],
    )
)

chart = (base_chart + desv_chart).properties(height=340)
st.altair_chart(chart, use_container_width=True)

# -----------------------------
# Movimientos — PRON
# -----------------------------
st.subheader("Movimientos (formato tesorería) — PRON (pendientes)")

mov_pron = view_pron.copy()
mov_pron["VTO. PAGO"] = mov_pron["FECHA"].dt.strftime("%d-%m-%y")
mov_pron["COBRADO/PAGADO"] = mov_pron["ESTATUS"]

pron_cols = ["VTO. PAGO", "CONCEPTO", "CLIENTE", "COBRADO/PAGADO", "COBROS", "PAGOS", "SALDO"]
for c in pron_cols:
    if c not in mov_pron.columns:
        mov_pron[c] = ""

mov_pron_out = mov_pron[pron_cols].copy()

styled_pron = (
    mov_pron_out.style
    .applymap(color_saldo, subset=["SALDO"])
    .applymap(style_estado, subset=["COBRADO/PAGADO"])
    .format({"COBROS": eur, "PAGOS": eur, "SALDO": eur})
)
st.dataframe(styled_pron, use_container_width=True)

# -----------------------------
# Movimientos — REAL
# -----------------------------
st.subheader("Movimientos (formato tesorería) — REAL (estimado/ejecutado)")

mov_real = view_real.copy()
mov_real["VTO. PAGO"] = mov_real["FECHA"].dt.strftime("%d-%m-%y")
mov_real["COBRADO/PAGADO"] = mov_real["ESTATUS"]

real_cols = ["VTO. PAGO", "CONCEPTO", "CLIENTE", "COBRADO/PAGADO", "COBROS", "PAGOS", "SALDO"]
for c in real_cols:
    if c not in mov_real.columns:
        mov_real[c] = ""

mov_real_out = mov_real[real_cols].copy()

styled_real = (
    mov_real_out.style
    .applymap(color_saldo, subset=["SALDO"])
    .applymap(style_estado, subset=["COBRADO/PAGADO"])
    .format({"COBROS": eur, "PAGOS": eur, "SALDO": eur})
)
st.dataframe(styled_real, use_container_width=True)

# -----------------------------
# Resumen mensual
# -----------------------------
st.subheader("Resumen mensual")

modo_resumen = st.radio(
    "Vista resumen mensual",
    options=["REAL", "PRON", "AMBOS"],
    index=0,
    horizontal=True
)

def resumen_mensual(df_base: pd.DataFrame, d_from: date, d_to: date) -> pd.DataFrame:
    df = df_base.copy()
    df = df[(df["FECHA"].dt.date >= d_from) & (df["FECHA"].dt.date <= d_to)].copy()
    df["MES"] = df["FECHA"].dt.to_period("M").astype(str)

    monthly_visible = df.groupby("MES", as_index=False).agg(
        COBROS=("COBROS", "sum"),
        PAGOS=("PAGOS", "sum"),
        NETO=("NETO", "sum"),
    )
    monthly_close = df.groupby("MES", as_index=False).agg(
        SALDO_CIERRE=("SALDO", "last")
    )
    return monthly_visible.merge(monthly_close, on="MES", how="left")

monthly_pron = resumen_mensual(consolidado_pron2, d_from, d_to)
monthly_real = resumen_mensual(consolidado_real2, d_from, d_to)

def show_monthly_table(df: pd.DataFrame, titulo: str):
    st.markdown(f"#### {titulo}")
    styled = df.style.format({c: eur for c in df.columns if c != "MES"})
    st.dataframe(styled, use_container_width=True)

if modo_resumen == "REAL":
    show_monthly_table(monthly_real, "Resumen mensual — REAL (estimado/ejecutado)")
elif modo_resumen == "PRON":
    show_monthly_table(monthly_pron, "Resumen mensual — PRON (pendientes)")
else:
    show_monthly_table(monthly_real, "Resumen mensual — REAL (estimado/ejecutado)")
    show_monthly_table(monthly_pron, "Resumen mensual — PRON (pendientes)")

# -----------------------------
# Exportar a Excel
# -----------------------------
st.subheader("Exportar (lo visible)")

export_mov_pron = mov_pron_out.copy()
export_mov_real = mov_real_out.copy()

for df_exp in (export_mov_pron, export_mov_real):
    for c in ["COBROS", "PAGOS", "SALDO"]:
        df_exp[c] = pd.to_numeric(df_exp[c], errors="coerce")

export_month_pron = monthly_pron.copy()
export_month_real = monthly_real.copy()
for df_exp in (export_month_pron, export_month_real):
    for c in ["COBROS", "PAGOS", "NETO", "SALDO_CIERRE"]:
        df_exp[c] = pd.to_numeric(df_exp[c], errors="coerce")

def build_excel_bytes(
    df_mov_pron: pd.DataFrame,
    df_mov_real: pd.DataFrame,
    df_month_pron: pd.DataFrame,
    df_month_real: pd.DataFrame
) -> bytes:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df_mov_pron.to_excel(writer, index=False, sheet_name="Movimientos_PRON")
        df_mov_real.to_excel(writer, index=False, sheet_name="Movimientos_REAL")
        df_month_pron.to_excel(writer, index=False, sheet_name="Resumen_PRON")
        df_month_real.to_excel(writer, index=False, sheet_name="Resumen_REAL")
    bio.seek(0)
    return bio.read()

xlsx_bytes = build_excel_bytes(export_mov_pron, export_mov_real, export_month_pron, export_month_real)

st.download_button(
    "Descargar Excel (XLSX)",
    data=xlsx_bytes,
    file_name="tesoreria_export.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
